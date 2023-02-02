import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook

#index = df['Index']
#date = df['Date']
#time = df['Time']
#operator = df['Operator']
#company = df['Company']
#bbls = df['BBLS']

df = pd.read_excel('Briscoe Carla Ranch 17 HA HB HC HD HE HF HG Trucking Sheet (version 1).xlsx')

df = df.rename(columns={
    df.columns[0]:'Index', 
    df.columns[1]:'Date', 
    df.columns[2]:'Time',
    df.columns[3]:'Operator',
    df.columns[4]:'Company',
    df.columns[5]:'Driver',
    df.columns[6]:'BBLS'})

df['Company'] = df['Company'].str.lower()

wb = load_workbook('Trucks Committed.xlsx')
ws_committed = wb['Committed']
ws_individual =  wb['Individuals']

start = int(input('Starting Range ')) + 4
end = int(input('Ending Range ')) + 4
shift = input('Enter shift: (Day or Night)\n')

col_wanted = df.loc[start:end, ['Index', 'Date', 'Time', 'Operator', 'Company', 'Driver','BBLS']]
#names_wanted = df.loc[start:end, ['Driver']]

if shift == 'Day' or shift == 'day':    

    #Gets all Tidal within selected range and sums the total bbls
    tidal = col_wanted[(df['Company'] == 'tidal')]
    tidal_bbls_total = tidal.loc[:, ['BBLS']].sum(0)
    
    select = col_wanted[(df['Company'] == 'select/tidal')]
    select_bbls_total = select.loc[:, ['BBLS']].sum(0)

    combined_ST_total = tidal_bbls_total + select_bbls_total

    #Gets Drivers
    tidal_drivers = tidal.loc[:, ['Driver']]
    select_drivers = select.loc[:, ['Driver']]

    #Gets the number of loads from select
    number_of_tidal_loads = len(col_wanted[df['Company'] == 'tidal'])
    number_of_select_loads = len(col_wanted[df['Company'] == 'select/tidal'])

    combined_ST_loads = number_of_tidal_loads + number_of_select_loads

    
    
    ##########################################################################
    #Gets all R&L within selected range and sums the total bbls
    rl = col_wanted[(df['Company'] == 'r&l')]
    rl_bbls_total = rl.loc[:, ['BBLS']].sum(0)

    number_of_rl_loads = len(col_wanted[df['Company'] == 'r&l'])

    ##########################################################################
    #Gets all WES within selected range and sums the total bbls
    wes = col_wanted[(df['Company'] == 'wes')]
    wes_bbls_total = wes.loc[:, ['BBLS']].sum(0)

    number_of_wes_loads = len(col_wanted[df['Company'] == 'wes'])
    
    ##########################################################################
    #Gets all One within selected range and sums the total bbls
    one = col_wanted[(df['Company'] == 'one')]
    one_bbls_total = one.loc[:, ['BBLS']].sum(0)

    number_of_one_loads = len(col_wanted[df['Company'] == 'one'])

    ##########################################################################    
    #Gets all Knowles within selected range and sums the total bbls
    knowles = col_wanted[(df['Company'] == 'knowles')] 
    knowles_bbls_total = knowles.loc[:, ['BBLS']].sum(0)

    number_of_knowles_loads = len(col_wanted[df['Company'] == 'knowles'])
    
    ##########################################################################    
    #Gets all Kasper within selected range and sums the total bbls
    kasper = col_wanted[(df['Company'] == 'kasper')]
    kasper_bbls_total = kasper.loc[:, ['BBLS']].sum(0)
    
    number_of_Kasper_loads = len(col_wanted[df['Company'] == 'kasper'])

    ##########################################################################    
    #Gets all Pro Field within selected range and sums the total bbls
    pro_field = col_wanted[(df['Company'] == 'pro field')]
    pro_field_bbls_total = pro_field.loc[:, ['BBLS']].sum(0)

    number_of_pro_field_loads = len(col_wanted[df['Company'] == 'pro field'])

    ##########################################################################
    #Gets all Mouflon within selected range and sums the total bbls
    mouflon = col_wanted[(df['Company'] == 'mouflon')]
    mouflon_bbls_total = mouflon.loc[:, ['BBLS']].sum(0)

    mf_fi = col_wanted[(df['Company'] == 'finaly/mouflon')]
    mf_fi_bbls_total = mf_fi.loc[:, ['BBLS']].sum(0)

    mf_tc = col_wanted[(df['Company'] == 'techwater/mouflon')]
    mf_tc_bbls_total = mf_tc.loc[:, ['BBLS']].sum(0)

    combined_mf_total = mouflon_bbls_total + mf_fi_bbls_total + mf_tc_bbls_total

    number_of_mouflon_loads = len(col_wanted[df['Company'] == 'mouflon'])
    number_of_mf_fi__loads = len(col_wanted[df['Company'] == 'finaly/mouflon'])
    number_of_mf_tc_loads = len(col_wanted[df['Company'] == 'techwater/mouflon'])

    combined_mf_fi_loads = number_of_mouflon_loads + number_of_mf_fi__loads + number_of_mf_tc_loads

    ##########################################################################    
    #Gets all Pro Field within selected range and sums the total bbls
    clarot = col_wanted[(df['Company'] == 'clarot')]
    clarot_bbls_total = clarot.loc[:, ['BBLS']].sum(0)

    number_of_clarot_loads = len(col_wanted[df['Company'] == 'clarot'])
    
    ##########################################################################    
    #iterates (inputs) all the gattered values into a specific column and row
    company_bbls_total = [
        int(combined_ST_total),
        int(rl_bbls_total), 
        int(wes_bbls_total), 
        int(one_bbls_total),
        int(combined_mf_total),
        int(kasper_bbls_total),
        int(pro_field_bbls_total),
        int(knowles_bbls_total),
        int(clarot_bbls_total)]

    for i, value in enumerate(company_bbls_total):
        ws_committed.cell(row=i+4, column=6, value=value)

    company_loads_total = [
        int(combined_ST_loads),
        int(number_of_rl_loads),
        int(number_of_wes_loads),
        int(number_of_one_loads),
        int(combined_mf_fi_loads),
        int(number_of_Kasper_loads),
        int(number_of_pro_field_loads),
        int(number_of_knowles_loads),
        int(number_of_clarot_loads)]

    for i, value in enumerate(company_loads_total):
        ws_committed.cell(row=i+4, column=5, value=value)

    driver_list = [
        tidal_drivers,
        select_drivers
    ]   

    for driver in enumerate(driver_list):
        ws_individual.cell(row=1, column=1)

    wb.save('Trucks Committed.xlsx')
'''''
if shift == 'Night' or shift == 'night':
    #Gets all Tidal within selected range and sums the total bbls
    tidal = col_wanted[(df['Company'] == 'tidal') & (df['Operator'] == 'Dago N')]
    tidal_bbls_total = tidal.loc[:, ['BBLS']].sum(0)
    
    select = col_wanted[(df['Company'] == 'select/tidal') & (df['Operator'] == 'Dago N')]
    select_bbls_total = select.loc[:, ['BBLS']].sum(0)

    combined_ST_total = tidal_bbls_total + select_bbls_total

    #Gets the number of loads from select
    number_of_tidal_loads = len(col_wanted[df['Company'] == 'tidal'])
    number_of_select_loads = len(col_wanted[df['Company'] == 'select/tidal'])

    combined_ST_loads = number_of_tidal_loads + number_of_select_loads

    ##########################################################################
    #Gets all R&L within selected range and sums the total bbls
    rl = col_wanted[(df['Company'] == 'r&l') & (df['Operator'] == 'Dago N')]
    rl_bbls_total = rl.loc[:, ['BBLS']].sum(0)

    number_of_rl_loads = len(col_wanted[df['Company'] == 'r&l'])

    ##########################################################################
    #Gets all WES within selected range and sums the total bbls
    wes = col_wanted[(df['Company'] == 'wes') & (df['Operator'] == 'Dago N')]
    wes_bbls_total = wes.loc[:, ['BBLS']].sum(0)

    number_of_wes_loads = len(col_wanted[df['Company'] == 'wes'])
    
    ##########################################################################
    #Gets all One within selected range and sums the total bbls
    one = col_wanted[(df['Company'] == 'one') & (df['Operator'] == 'Dago N')]
    one_bbls_total = one.loc[:, ['BBLS']].sum(0)

    number_of_one_loads = len(col_wanted[df['Company'] == 'one'])

    ##########################################################################    
    #Gets all Knowles within selected range and sums the total bbls
    knowles = col_wanted[(df['Company'] == 'knowles') & (df['Operator'] == 'Dago N')] 
    knowles_bbls_total = knowles.loc[:, ['BBLS']].sum(0)

    number_of_knowles_loads = len(col_wanted[df['Company'] == 'knowles'])
    
    ##########################################################################    
    #Gets all Kasper within selected range and sums the total bbls
    kasper = col_wanted[(df['Company'] == 'kasper') & (df['Operator'] == 'Dago N')]
    kasper_bbls_total = kasper.loc[:, ['BBLS']].sum(0)
    
    number_of_Kasper_loads = len(col_wanted[df['Company'] == 'kasper'])

    ##########################################################################    
    #Gets all Pro Field within selected range and sums the total bbls
    pro_field = col_wanted[(df['Company'] == 'pro field') & (df['Operator'] == 'Dago N')]
    pro_field_bbls_total = pro_field.loc[:, ['BBLS']].sum(0)

    number_of_pro_field_loads = len(col_wanted[df['Company'] == 'pro field'])

    ##########################################################################
    #Gets all Mouflon within selected range and sums the total bbls
    mouflon = col_wanted[(df['Company'] == 'mouflon') & (df['Operator'] == 'Dago N')]
    mouflon_bbls_total = mouflon.loc[:, ['BBLS']].sum(0)

    mf_fi = col_wanted[(df['Company'] == 'finaly/mouflon') & (df['Operator'] == 'Dago N')]
    mf_fi_bbls_total = mf_fi.loc[:, ['BBLS']].sum(0)

    mf_tc = col_wanted[(df['Company'] == 'techwater/mouflon') & (df['Operator'] == 'Dago N ')]
    mf_tc_bbls_total = mf_tc.loc[:, ['BBLS']].sum(0)

    combined_mf_total = mouflon_bbls_total + mf_fi_bbls_total + mf_tc_bbls_total

    number_of_mouflon_loads = len(col_wanted[df['Company'] == 'mouflon'])
    number_of_mf_fi__loads = len(col_wanted[df['Company'] == 'finaly/mouflon'])
    number_of_mf_tc_loads = len(col_wanted[df['Company'] == 'techwater/mouflon'])

    combined_mf_fi_loads = number_of_mouflon_loads + number_of_mf_fi__loads + number_of_mf_tc_loads
    
    ##########################################################################    
    #Gets all Pro Field within selected range and sums the total bbls
    clarot = col_wanted[(df['Company'] == 'clarot') & (df['Operator'] == 'Dago N')]
    clarot_bbls_total = clarot.loc[:, ['BBLS']].sum(0)

    number_of_clarot_loads = len(col_wanted[df['Company'] == 'clarot'])

    ##########################################################################    
    #iterates (inputs) all the gattered values into a specific column and row
    company_bbls_total = [
        int(combined_ST_total),
        int(rl_bbls_total), 
        int(wes_bbls_total), 
        int(one_bbls_total),
        int(combined_mf_total),
        int(kasper_bbls_total),
        int(pro_field_bbls_total),
        int(knowles_bbls_total),
        int(clarot_bbls_total)]

    for i, value in enumerate(company_bbls_total):
        ws_committed.cell(row=i+23, column=6, value=value)

    company_loads_total = [
        int(combined_ST_loads),
        int(number_of_rl_loads),
        int(number_of_wes_loads),
        int(number_of_one_loads),
        int(combined_mf_fi_loads),
        int(number_of_Kasper_loads),
        int(number_of_pro_field_loads),
        int(number_of_knowles_loads),
        int(number_of_clarot_loads)]

    for i, value in enumerate(company_loads_total):
        ws_committed.cell(row=i+23, column=5, value=value)    

    wb.save('Trucks Committed.xlsx')
    '''''

#print(ws_committed)
#print(names_wanted)
print(col_wanted)
print(tidal)
print(len(tidal_drivers))
print(driver_list)
#print(df['Company'].head(10))
#print(df.columns)