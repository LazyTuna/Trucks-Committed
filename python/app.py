#Program Created By Jose A Hernandez Jr, started on 12/2022, Beta Version 1
#This program grabs the selected trucking report and calculates the number of loads and bbls.
#Sending it to the selected committed trucks.

import PySimpleGUI as sg
import pandas as pd
from openpyxl import Workbook, load_workbook
import numpy as np
import os
from datetime import datetime, timedelta

desktop_path = os.path.expanduser('~') #  path.expanduser('~') gets the home directory

sg.theme('DarkGreen6')
operator_name = []
truck_company = []

layout = [
    [
        sg.Text('Trucking Report:'), 
        sg.Push(), sg.InputText(key='-TRUCKING_REPORT-'), sg.FileBrowse(initial_folder=desktop_path, file_types = [('Excel Files', '*.xlsx')])
    ],
    [
        sg.Text('Trucks Committed:'),
        sg.Push(), sg.InputText(key='-TRUCKS_COMMITTED-'), sg.FileBrowse(initial_folder=desktop_path, file_types = [('Excel Files', '*.xlsx')])
    ],
    [
        sg.Text('Starting Range:'), sg.InputText(key='-STARTING_RANGE-', size=(6,1)), 
        sg.Text('Ending Range:'), sg.InputText(key='-ENDING_RANGE-', size=(6,1)),
        sg.Text('Day or Night'), sg.InputText(key='-DAY_NIGHT-', size=(6,1))
    ],
    [
        #sg.Text('Operator:'),
        #sg.Combo(operator_name, size=(10,1), key='-OPERATOR-', bind_return_key=True),
        sg.Text('Truck Comp:'),
        sg.Combo(truck_company, size=(25,1), key='-COMPANY-', bind_return_key=True),
    ],
    [sg.Button('Submit'), sg.Button('Update'), sg.Button('Instructions'),sg.Exit()]
]

window = sg.Window('Committed Trucks', layout, finalize=True)

def instruction():
    instruction = [
        [sg.Text('1.) This programs calculates total loads and bbls for the current shift.')],
        [sg.Text('2.) Choose the desired Trucking Sheet excel file.')],
        [sg.Text('3.) Choose the Trucks Committed excel file, should be included with app.')],
        [sg.Text('4.) Starting Range will be \'Column 1\' number from the start of your shift on the Trucking Sheet.')],
        [sg.Text('5.) Ending Range will be \'Column 1\' number from the end of your shift on the Trucking Sheet.')],
        [sg.Text('6.) Day or Night with be the shift you are on \'day\' or \'night.\'')],
        [sg.Text('7.) To use the \'Trucking Comp:\' section, type in the trucking company exactly as you typed it')],
        [sg.Text('      on the excel file and press \'Enter\'. This will add the company to the list.')],
        [sg.Text('      To check if it was added, press the drop down arrow.')],
        [sg.Text('      If satisfied with the list, press the \'Update\' button.')],
        [sg.Text('******You can leave as many companys on the list even if they dont show up that day.******')],
        [sg.Text('8.) To remove a Trucking Company from the list, click the drop down menu, select the company')],
        [sg.Text('      and press \'Enter\'. To check if it was removed, press the drop down arrow.')],
        [sg.Text('      If satisfied with the list, press the \'Update\' button.')],
        [sg.Text('9.) When all fields have been enterted press the \'Submit\' button. App should stay open.')],
        [sg.Text('*****If app closes, and error has occurred. Make sure both excel files are closed and try again.*****.')],
    ]

    instruction_window = sg.Window('Instructions', instruction, modal=True)

    while True:
        event, values = instruction_window.read()
        if event in (sg.WIN_CLOSED, 'Exit'):
            break

        window.close()

sg.user_settings_filename(path='.')
ops_list = sg.user_settings_get_entry('combo list')
truck_company_list = sg.user_settings_get_entry('combo list2')

#if ops_list != None:
#    operator_name = ops_list
#    window['-OPERATOR-'].update(values=operator_name)

if truck_company != None:
    truck_company = truck_company_list
    window['-COMPANY-'].update(values=truck_company)

while True:
    event, values = window.read()

    #if event == '-OPERATOR-':
    #    print(values['-OPERATOR-'])
    #    print(operator_name)
    #    if values['-OPERATOR-'] not in operator_name: #this
    #        operator_name.append(values['-OPERATOR-'])
    #        window['-OPERATOR-'].update(values=operator_name,value=values['-OPERATOR-'])
    #    else: #this line deletes str when selected and pressed enter
    #        operator_name.remove(values['-OPERATOR-'])
    #        window['-OPERATOR-'].update(values=operator_name,value='')
    
    if event == '-COMPANY-':
        print(values['-COMPANY-'])
        if values['-COMPANY-'] not in truck_company:
            truck_company.append(values['-COMPANY-'])
            window['-COMPANY-'].update(values=truck_company,value=values['-COMPANY-'])
        else:
            truck_company.remove(values['-COMPANY-'])
            window['-COMPANY-'].update(values=truck_company,value='')

    if event == 'Submit':
        trucking_report = values['-TRUCKING_REPORT-']
        trucks_committed = values['-TRUCKS_COMMITTED-']
        start_sg = values['-STARTING_RANGE-']
        end_sg = values['-ENDING_RANGE-']
        shift_sg = values['-DAY_NIGHT-']
        #operator = values['-OPERATOR-']


        df = pd.read_excel(trucking_report)
        df2 = pd.read_excel(trucks_committed)

        df = df.rename(columns={
                df.columns[0]:'Index', 
                df.columns[1]:'Date', 
                df.columns[2]:'Time',
                df.columns[3]:'Operator',
                df.columns[4]:'Company',
                df.columns[5]:'Driver',
                df.columns[6]:'BBLS'})
        
        df['Company'] = df['Company'].str.lower()
        df['Time'] = df['Time'].str.lower()


        wb = load_workbook(trucks_committed)
        ws_committed = wb["Committed"]
        ws_individual =  wb['Drivers']

        start = int(start_sg) + 4
        end = int(end_sg) + 4
        shift = shift_sg.lower()

        col_wanted = df.loc[start:end, ['Index', 'Date', 'Time', 'Operator', 'Company', 'Driver','BBLS']]

        truck_company_list = [item.lower() for item in truck_company_list]
        bbls_list = []
        loads_list = []
        temp_list = []

        driver_list = []
        driver_list_temp = []
        driver_bbls_list = []
        driver_loads_list = []

        time_list = []
        time_list_object = []
        time_list_td = []
        total_time = timedelta
##///////////////////////////////////////////////////////////////////////////////////////////////
        if shift == 'day':
            #this line deletes any values in the wb so you wont have to
            for row in ws_committed['A2:F16']:
                for cell in row:
                    cell.value = None

            for row in ws_committed['A26:F40']:
                for cell in row:
                    cell.value = None 

            for row in ws_individual['A2:D25']:
                for cell in row:
                    cell.value = None            

            ### This section compares the trucks inputed in the app and on the excel file
            ### if it matches, will input it to the appropriate trucks committed section
            for comp in truck_company_list:
                if comp in col_wanted[(df['Company'] == comp)].values:

                    temp_list.append(comp.title())
                    truck = col_wanted[(df['Company'] == comp)]
                    bbl = int(truck.loc[:, ['BBLS']].sum(0))
                    loads = int(len(col_wanted[df['Company'] == comp]))
                    bbls_list.append(bbl)
                    loads_list.append(loads)
                    
                    for i, value in enumerate(temp_list):
                        ws_committed.cell(row=i+2, column=2, value=value)  

                    for i, value in enumerate(loads_list):
                        ws_committed.cell(row=i+2, column=5, value=value) 

                    for i, value in enumerate(bbls_list):
                        ws_committed.cell(row=i+2, column=6, value=value)

            ### This section will get the drivers names 
            ###
            drivers = df.loc[start:end,['Driver']].values
            for driver in drivers:
                driver_temp = str(driver).replace("['",'').replace("']", '')
                driver_list.append(driver_temp)

                #remove duplicates from list
                driver_list_temp = set(driver_list)
                driver_list = list(driver_list_temp)

            for i, value in enumerate(driver_list):
                ws_individual.cell(row=i+2, column=1, value=str(value))

            ### This section will get the total loads and total bbls for each driver for the day
            ###
            for driver in driver_list:
                if driver in col_wanted[(df['Driver'] == driver)].values:
                    driver_col = col_wanted[(df['Driver'] == driver)]
                    driver_bbl = int(driver_col.loc[:, ['BBLS']].sum(0))
                    driver_loads = int(len(col_wanted[df['Driver'] == driver]))
                    driver_bbls_list.append(driver_bbl)
                    driver_loads_list.append(driver_loads)

            for i, value in enumerate(driver_loads_list):
                ws_individual.cell(row=i+2, column=2, value=value)

            for i, value in enumerate(driver_bbls_list):
                ws_individual.cell(row=i+2, column=3, value=value)

            ### This will get Time
            times = df.loc[start:end,['Time']].values
            for time in times:
                time_temp = str(time).replace("am",'').replace("pm",'').replace(' am','').replace(' pm', '').replace("['",'').replace("']", '')
                time_list.append(time_temp)

            #gets the time string and converts it to an object
            for time_string in time_list:
                time_obj = datetime.strptime(time_string, "%H:%M").time()
                time_list_object.append(time_obj)

            for time_obj in time_list_object:
                td = timedelta(hours=time_obj.hour, minutes=time_obj.minute, seconds=time_obj.second)
                time_list_td.append(td)
                #total_time += td
                print(time_list_td)
                print(td)
                #print(total_time)
                

            #days,seconds = divmod(total_time.seconds, 86400)
            #total_time = datetime.time(seconds//3600, (seconds//60)%60, seconds%60)
            #print(total_time)

            wb.save(trucks_committed)

##///////////////////////////////////////////////////////////////////////////////////////////////
        if shift == 'night':
            for row in ws_committed['A26:F40']:
                for cell in row:
                    cell.value = None

            for row in ws_individual['F2:I25']:
                for cell in row:
                    cell.value = None  

            ### This section compares the trucks inputed in the app and on the excel file
            ### if it matches, will input it to the appropriate trucks committed section
            for comp in truck_company_list:
                if comp in col_wanted[(df['Company'] == comp)].values:

                    temp_list.append(comp.title())
                    truck = col_wanted[(df['Company'] == comp)]
                    bbl = int(truck.loc[:, ['BBLS']].sum(0))
                    loads = int(len(col_wanted[df['Company'] == comp]))
                    bbls_list.append(bbl)
                    loads_list.append(loads)
                    
                    for i, value in enumerate(temp_list):
                        ws_committed.cell(row=i+26, column=2, value=value)

                    for i, value in enumerate(loads_list):
                        ws_committed.cell(row=i+26, column=5, value=value) 

                    for i, value in enumerate(bbls_list):
                        ws_committed.cell(row=i+26, column=6, value=value)

            ### This section will get the drivers names 
            ###
            drivers = df.loc[start:end,['Driver']].values
            for driver in drivers:
                driver_temp = str(driver).replace("['",'').replace("']", '')
                driver_list.append(driver_temp.title())

                #remove duplicates from list
                driver_list_temp = set(driver_list)
                driver_list = list(driver_list_temp)

            for i, value in enumerate(driver_list):
                ws_individual.cell(row=i+2, column=6, value=str(value))

            ### This section will get the total loads and total bbls for each driver for the day
            ###
            for driver in driver_list:
                if driver in col_wanted[(df['Driver'] == driver)].values:
                    
                    driver_col = col_wanted[(df['Driver'] == driver)]
                    driver_bbl = int(driver_col.loc[:, ['BBLS']].sum(0))
                    driver_loads = int(len(col_wanted[df['Driver'] == driver]))
                    driver_bbls_list.append(driver_bbl)
                    driver_loads_list.append(driver_loads)

            for i, value in enumerate(driver_loads_list):
                ws_individual.cell(row=i+2, column=7, value=value)

            for i, value in enumerate(driver_bbls_list):
                ws_individual.cell(row=i+2, column=8, value=value)
                        
            wb.save(trucks_committed)
        
        print(col_wanted)

    if event == 'Update':
        sg.user_settings_set_entry('combo list', operator_name)
        sg.user_settings_set_entry('combo list2', truck_company)
        print(truck_company_list)

    if event == 'Instructions':
        instruction()

    if event in (sg.WIN_CLOSED, 'Exit'):
        break

window.close()